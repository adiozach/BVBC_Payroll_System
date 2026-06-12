"""
BVBC Payroll – Excel Export Module  |  v2.0 with Official Logo
"""
import os
from datetime import datetime

REPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
LOGO_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "logo.png")
os.makedirs(REPORT_DIR, exist_ok=True)

MONTHS = ["","January","February","March","April","May","June",
          "July","August","September","October","November","December"]

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

def _require():
    if not OPENPYXL_OK:
        raise RuntimeError(
            "openpyxl is not installed.\n\n"
            "Fix: Open Command Prompt and run:\n"
            "  pip install openpyxl\n\n"
            "Then restart the application.\n"
            "Use install_and_run.bat to start correctly.")

# Colours (no #)
MAR2="4A0909"; MAR1="6D0E0E"; GOLD="C9A84C"
LGOLD="FFF8E7"; WHT="FFFFFF"; DTXT="1A0A0A"

def _b(): 
    s = Side(style="thin", color="D4B483")
    return Border(left=s, right=s, top=s, bottom=s)
def _f(h):  return PatternFill("solid", fgColor=h)
def _ft(bold=False, color=DTXT, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)
def _al(h="center", v="center", wrap=False, ind=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=ind)

def _hrow(ws, row, vals, bg=MAR1, fg=WHT, sz=10, ht=24):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = _ft(True, fg, sz); cell.fill = _f(bg)
        cell.alignment = _al(wrap=True); cell.border = _b()
    ws.row_dimensions[row].height = ht

def _dcell(ws, r, c, val, money=False, alt=False, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _ft(bold=bold); cell.fill = _f(LGOLD if alt else WHT)
    cell.border = _b()
    if money and isinstance(val, (int, float)):
        cell.number_format = "#,##0.00"; cell.alignment = _al("right")
    elif isinstance(val, (int, float)):
        cell.alignment = _al("center")
    else:
        cell.alignment = _al("left", ind=1)

def _gold_bar(ws, row, ncols):
    for c in range(1, ncols+1):
        ws.cell(row=row, column=c).fill = _f(GOLD)
    ws.row_dimensions[row].height = 4

def _embed_logo(ws, cell_ref="A1", w_px=60, h_px=60):
    """Embed the school logo if available."""
    if OPENPYXL_OK and os.path.exists(LOGO_PATH):
        try:
            xl_img = XLImage(LOGO_PATH)
            xl_img.width  = w_px
            xl_img.height = h_px
            ws.add_image(xl_img, cell_ref)
        except Exception:
            pass

def _title_block(ws, ncols, title, subtitle):
    """School header with logo, name and report title."""
    # Row 1 – school name (leave space for logo in col A)
    ws.merge_cells(f"B1:{get_column_letter(ncols)}1")
    c = ws["B1"]; c.value = "BAPTIST VOICE BIBLE COLLEGE"
    c.font = _ft(True, GOLD, 15); c.fill = _f(MAR2)
    c.alignment = _al(); ws.row_dimensions[1].height = 42

    # Fill logo column same bg
    ws["A1"].fill = _f(MAR2)

    # Row 2 – report title
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]; c.value = title
    c.font = _ft(True, WHT, 12); c.fill = _f(MAR1)
    c.alignment = _al(); ws.row_dimensions[2].height = 24

    # Row 3 – subtitle / date
    ws.merge_cells(f"A3:{get_column_letter(ncols)}3")
    c = ws["A3"]; c.value = subtitle
    c.font = _ft(False, "6D4C2E", 9); c.fill = _f(LGOLD)
    c.alignment = _al(); ws.row_dimensions[3].height = 16

    _gold_bar(ws, 4, ncols)
    _embed_logo(ws, "A1", w_px=58, h_px=58)


# ═══════════════════════════════════════════════════════════════════════════
def export_payroll_excel(records, month, year):
    _require()
    wb = Workbook(); ws = wb.active
    ws.title = f"Payroll {MONTHS[month]} {year}"
    ws.sheet_view.showGridLines = False
    NCOLS = 19

    _title_block(ws, NCOLS,
                 f"MONTHLY PAYROLL REPORT  —  {MONTHS[month].upper()} {year}",
                 f"Generated: {datetime.now().strftime('%B %d, %Y  %I:%M %p')}")

    headers = ["Emp ID","Employee Name","Department","Days",
               "Basic Pay","OT Hrs","OT Pay","Gross Pay",
               "SSS","PhilHealth","Pag-IBIG","Cash Adv",
               "HDMF Loan","SSS Loan","Other Ded","Late Ded",
               "Absent Ded","Total Ded","Net Pay"]
    _hrow(ws, 5, headers, ht=28)
    ws.freeze_panes = "A6"

    for i, rec in enumerate(records):
        r = i + 6; alt = i % 2 == 1
        row_data = [
            (rec["employee_id"],             False),
            (f"{rec['last_name']}, {rec['first_name']}", False),
            (rec["department"] or "",        False),
            (float(rec["days_worked"]),      False),
            (rec["basic_pay"],               True),
            (float(rec["overtime_hours"]),   False),
            (rec["overtime_pay"],            True),
            (rec["gross_pay"],               True),
            (rec["sss"],                     True),
            (rec["philhealth"],              True),
            (rec["pagibig"],                 True),
            (rec["cash_advance"],            True),
            (rec["hdmf_loan"],               True),
            (rec["sss_loan"],                True),
            (rec["other_deductions"],        True),
            (rec["late_deduction"],          True),
            (rec["absent_deduction"],        True),
            (rec["total_deductions"],        True),
            (rec["net_pay"],                 True),
        ]
        for col, (val, money) in enumerate(row_data, 1):
            _dcell(ws, r, col, val, money=money, alt=alt, bold=(col==19))

    # Totals
    tr = len(records) + 6
    for col in range(1, NCOLS+1):
        c = ws.cell(row=tr, column=col)
        c.fill = _f(GOLD); c.font = _ft(True, MAR2)
        c.border = _b(); c.alignment = _al()
    ws.cell(row=tr, column=1).value = "GRAND TOTAL"
    for col in [5,7,8,9,10,11,12,13,14,15,16,17,18,19]:
        L = get_column_letter(col)
        c = ws.cell(row=tr, column=col)
        if records: c.value = f"=SUM({L}6:{L}{tr-1})"
        c.number_format = "#,##0.00"; c.alignment = _al("right")

    widths = [10,24,16,8,13,7,11,13,10,12,10,10,11,10,10,10,11,11,13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.page_setup.orientation = "landscape"; ws.page_setup.paperSize = 9
    fname = f"Payroll_{MONTHS[month]}_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path  = os.path.join(REPORT_DIR, fname)
    wb.save(path); return path


# ═══════════════════════════════════════════════════════════════════════════
def export_payslip_excel(rec, emp):
    """
    BVBC Payslip Excel export.
    Layout (all 4 columns A-D, no overlap):
      Rows  1-5   : Header (logo, title, period)
      Row   6     : EMPLOYEE INFORMATION header
      Rows  7-14  : Employee info (cols A-B merged label + value)
      Row  15     : Gold bar
      Row  16     : EARNINGS | DEDUCTIONS column headers
      Rows 17-21  : Earnings (cols A-B) + first 5 deductions (cols C-D)
      Rows 22+    : Remaining deductions (cols C-D only, cols A-B blank)
      Row  X      : SAVINGS header (cols C-D)
      Rows X+1..  : Savings rows (cols C-D)
      Row  Y      : NET PAY (full width A-D)
      Row  Y+2    : Signature line
    """
    _require()
    wb = Workbook(); ws = wb.active
    ws.title = "Payslip"; ws.sheet_view.showGridLines = False

    m_n  = int(rec["period_month"]); yr = rec["period_year"]
    ct   = rec["cutoff"]
    fn   = emp["first_name"]; ln = emp["last_name"]
    dept = emp.get("department") or ""; pos = emp.get("position") or ""
    ety  = emp.get("employment_type") or ""; eid = emp["employee_id"]

    def _v(k): return float(rec.get(k) or 0)

    def mhdr(rng, text, bg=MAR2, fg=GOLD, sz=12, bold=True, ht=28):
        ws.merge_cells(rng); sr = rng.split(":")[0]
        c = ws[sr]; c.value = text
        c.font = _ft(bold, fg, sz); c.fill = _f(bg); c.alignment = _al()
        rn = int("".join(x for x in sr if x.isdigit()))
        ws.row_dimensions[rn].height = ht

    def info_row(r, lbl, val, alt=False):
        """Full-width info row spanning A-B for label, C-D for value."""
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws[f"A{r}"]; lc.value = lbl
        lc.font = _ft(True, "6D4C2E", 9); lc.fill = _f(LGOLD if alt else WHT)
        lc.border = _b(); lc.alignment = _al("left", ind=1)
        ws.merge_cells(f"C{r}:D{r}")
        vc = ws[f"C{r}"]; vc.value = val
        vc.font = _ft(False, "1A0A0A", 9); vc.fill = _f(LGOLD if alt else WHT)
        vc.border = _b()
        if isinstance(val, (int, float)):
            vc.number_format = "#,##0.00"; vc.alignment = _al("right")
        else:
            vc.alignment = _al("left", ind=1)
        ws.row_dimensions[r].height = 20

    def earn_row(r, lbl, val, alt=False, bold=False):
        """Earnings: cols A (label) + B (value)."""
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = _ft(bold, "6D4C2E", 9); lc.fill = _f(LGOLD if alt else WHT)
        lc.border = _b(); lc.alignment = _al("left", ind=1)
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = _ft(bold, DTXT, 9); vc.fill = _f(LGOLD if alt else WHT)
        vc.border = _b()
        if isinstance(val, (int, float)):
            vc.number_format = "#,##0.00"; vc.alignment = _al("right")
        else:
            vc.alignment = _al("right")
        ws.row_dimensions[r].height = 20

    def blank_earn(r, alt=False):
        """Blank earnings cells to fill gap rows."""
        for col in [1, 2]:
            c = ws.cell(row=r, column=col)
            c.fill = _f(LGOLD if alt else WHT); c.border = _b()
        ws.row_dimensions[r].height = 20

    def ded_row(r, lbl, val, alt=False, bold=False, color=None):
        """Deduction: cols C (label) + D (value)."""
        bg = color if color else (LGOLD if alt else WHT)
        lc = ws.cell(row=r, column=3, value=lbl)
        lc.font = _ft(bold, "6D4C2E" if not bold else DTXT, 9)
        lc.fill = _f(bg); lc.border = _b(); lc.alignment = _al("left", ind=1)
        vc = ws.cell(row=r, column=4, value=val)
        vc.font = _ft(bold, DTXT, 9)
        vc.fill = _f(bg); vc.border = _b()
        if isinstance(val, (int, float)):
            vc.number_format = "#,##0.00"; vc.alignment = _al("right")
        else:
            vc.alignment = _al("right")
        ws.row_dimensions[r].height = 20

    def ded_header(r, text, bg_hex):
        ws.merge_cells(f"C{r}:D{r}")
        c = ws[f"C{r}"]; c.value = text
        c.font = _ft(True, WHT, 9); c.fill = _f(bg_hex)
        c.alignment = _al(); c.border = _b()
        ws.row_dimensions[r].height = 20

    # ── Rows 1-5: Header ────────────────────────────────────────────────────
    ws["A1"].fill = _f(MAR2)
    ws.merge_cells("B1:D1")
    c = ws["B1"]; c.value = "BAPTIST VOICE BIBLE COLLEGE"
    c.font = _ft(True, GOLD, 13); c.fill = _f(MAR2); c.alignment = _al()
    ws.row_dimensions[1].height = 48
    _embed_logo(ws, "A1", w_px=46, h_px=46)

    _gold_bar(ws, 2, 4)
    mhdr("A3:D3", "EMPLOYEE PAYSLIP", bg=MAR1, fg=WHT, sz=12, ht=26)
    mhdr("A4:D4", f"{MONTHS[m_n].upper()} {yr}  |  CUTOFF: {ct}",
         bg=LGOLD, fg=MAR2, sz=10, bold=False, ht=20)
    _gold_bar(ws, 5, 4)

    # ── Row 6: Employee Info header ─────────────────────────────────────────
    mhdr("A6:D6", "EMPLOYEE INFORMATION", bg=MAR1, fg=WHT, sz=10, ht=22)

    # ── Rows 7-14: Employee info ────────────────────────────────────────────
    rent_val = float(rec.get("rent") or emp.get("rent") or 0)
    gs_val   = max(float(emp.get("monthly_salary") or 0) - rent_val, 0)
    info_items = [
        ("Employee ID",     eid,                                    False),
        ("Full Name",       f"{fn} {ln}",                          True),
        ("Department",      dept,                                   False),
        ("Position",        pos,                                    True),
        ("Employment Type", ety,                                    False),
        ("Monthly Salary",  float(emp.get("monthly_salary") or 0), True),
        ("Rent",            rent_val,                               False),
        ("Gross Salary",    gs_val,                                 True),
    ]
    for i, (lbl, val, alt) in enumerate(info_items):
        info_row(7 + i, lbl, val, alt)

    # ── Row 15: Gold bar ────────────────────────────────────────────────────
    _gold_bar(ws, 15, 4)

    # ── Row 16: EARNINGS | DEDUCTIONS column headers ────────────────────────
    ws.merge_cells("A16:B16"); eh = ws["A16"]; eh.value = "EARNINGS"
    eh.font = _ft(True, WHT, 10); eh.fill = _f(MAR1)
    eh.alignment = _al(); eh.border = _b()
    ws.merge_cells("C16:D16"); dh = ws["C16"]; dh.value = "DEDUCTIONS"
    dh.font = _ft(True, WHT, 10); dh.fill = _f("8B1A1A")
    dh.alignment = _al(); dh.border = _b()
    ws.row_dimensions[16].height = 22

    # ── Earnings list ───────────────────────────────────────────────────────
    earn_items = [
        ("Days Worked",  float(rec["days_worked"]), False, False),
        ("Basic Pay",    rec["basic_pay"],           True,  False),
        ("Overtime Pay", rec["overtime_pay"],        False, False),
        (None,           None,                       True,  False),   # blank spacer
        ("GROSS PAY",    rec["gross_pay"],           False, True),
    ]

    # ── Deductions list ─────────────────────────────────────────────────────
    ded_items = [
        ("SSS Contribution", _v("sss"),              False, False),
        ("SSS WISP",         _v("sss_wisp"),         True,  False),
        ("SSS Loan",         _v("sss_loan"),         False, False),
        ("HDMF Con.",        _v("pagibig"),          True,  False),
        ("HDMF Loan",        _v("hdmf_loan"),        False, False),
        ("PhilHealth",       _v("philhealth"),       True,  False),
        ("COOP Loan",        _v("coop_loan"),        False, False),
        ("Alumni Fee",       _v("alumni_fee"),       True,  False),
        ("CA",               _v("cash_advance"),     False, False),
        ("Uniform",          _v("uniform"),          True,  False),
        ("Canteen",          _v("canteen"),          False, False),
        ("Late Deduction",   _v("late_deduction"),   True,  False),
        ("Absent Deduction", _v("absent_deduction"), False, False),
        ("Others",           _v("other_deductions"), True,  False),
        ("TOTAL DEDUCTIONS", rec["total_deductions"],False, True),
    ]

    # ── Savings list ────────────────────────────────────────────────────────
    sav_items = [
        ("COOP Savings",         _v("coop_savings"), False, False),
        ("Insurance",            _v("insurance"),    True,  False),
        ("Travel Fund",          _v("travel_fund"),  False, False),
        ("Sacrificial Offering", _v("sacrificial"),  True,  False),
        ("TOTAL SAVINGS",        _v("total_savings"),False, True),
    ]

    # Start row for data
    DATA_START = 17

    # Write earnings (cols A-B) — only as many rows as earn_items
    for i, (lbl, val, alt, bold) in enumerate(earn_items):
        r = DATA_START + i
        if lbl is None:
            blank_earn(r, alt)
        else:
            earn_row(r, lbl, val, alt, bold)

    # Fill blank earn cells for rows beyond earn length (so borders are consistent)
    earn_end_row = DATA_START + len(earn_items) - 1
    ded_end_row  = DATA_START + len(ded_items)  - 1
    sav_hdr_row  = ded_end_row + 1
    sav_end_row  = sav_hdr_row + len(sav_items)

    for r in range(earn_end_row + 1, sav_end_row + 2):
        blank_earn(r, alt=(r % 2 == 0))

    # Write deductions (cols C-D) — starting at DATA_START
    for i, (lbl, val, alt, bold) in enumerate(ded_items):
        r = DATA_START + i
        ded_row(r, lbl, val, alt, bold)

    # SAVINGS header
    ded_header(sav_hdr_row, "SAVINGS", "2D6A2D")

    # Write savings (cols C-D)
    for i, (lbl, val, alt, bold) in enumerate(sav_items):
        r = sav_hdr_row + 1 + i
        ded_row(r, lbl, val, alt, bold)

    # ── NET PAY row ─────────────────────────────────────────────────────────
    nr = sav_end_row + 2
    ws.merge_cells(f"A{nr}:C{nr}")
    nc = ws[f"A{nr}"]; nc.value = "NET PAY"
    nc.font = _ft(True, GOLD, 13); nc.fill = _f(MAR2)
    nc.alignment = _al("right", ind=2); nc.border = _b()
    vc = ws[f"D{nr}"]
    vc.value = rec["net_pay"]
    vc.font = _ft(True, GOLD, 13); vc.fill = _f(MAR2)
    vc.number_format = "#,##0.00"; vc.alignment = _al("right"); vc.border = _b()
    ws.row_dimensions[nr].height = 30

    # ── Signature line ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{nr+2}:D{nr+2}")
    sc = ws[f"A{nr+2}"]
    sc.value = "Received by: _______________________     Date: ____________"
    sc.font = _ft(False, "6D4C2E", 9); sc.alignment = _al()

    # ── Column widths ───────────────────────────────────────────────────────
    for col, w in zip("ABCD", [18, 14, 22, 14]):
        ws.column_dimensions[col].width = w

    ws.page_setup.paperSize = 9
    fname = f"Payslip_{eid}_{MONTHS[m_n]}_{yr}.xlsx"
    path  = os.path.join(REPORT_DIR, fname)
    wb.save(path); return path


# ═══════════════════════════════════════════════════════════════════════════
def export_deductions_excel(records, month, year):
    _require()
    wb=Workbook(); ws=wb.active
    ws.title=f"Deductions {MONTHS[month]} {year}"
    ws.sheet_view.showGridLines=False
    NCOLS=10
    _title_block(ws,NCOLS,
                 f"DEDUCTIONS REPORT  —  {MONTHS[month].upper()} {year}",
                 f"Generated: {datetime.now().strftime('%B %d, %Y  %I:%M %p')}")
    headers=["Emp ID","Name","SSS","PhilHealth","Pag-IBIG",
             "Cash Adv","HDMF Loan","SSS Loan","Other","Total Ded"]
    _hrow(ws,5,headers,ht=24)
    for i,rec in enumerate(records):
        r=i+6; alt=i%2==1
        row_data=[(rec["employee_id"],False),
                  (f"{rec['last_name']}, {rec['first_name']}",False),
                  (rec["sss"],True),(rec["philhealth"],True),(rec["pagibig"],True),
                  (rec["cash_advance"],True),(rec["hdmf_loan"],True),
                  (rec["sss_loan"],True),(rec["other_deductions"],True),
                  (rec["total_deductions"],True)]
        for col,(val,money) in enumerate(row_data,1):
            _dcell(ws,r,col,val,money=money,alt=alt,bold=(col==10))
    for col,w in zip("ABCDEFGHIJ",[10,22,10,12,10,10,10,10,10,12]):
        ws.column_dimensions[col].width=w
    fname=f"Deductions_{MONTHS[month]}_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path=os.path.join(REPORT_DIR,fname)
    wb.save(path); return path
